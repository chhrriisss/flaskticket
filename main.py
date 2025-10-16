from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
import json
import os
from datetime import datetime
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from functools import wraps
from models import db, User, FieldConfig, Package, EmailPreference
from flask_mail import Mail, Message
from threading import Thread

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///project.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = 'your-secret-key-change-this'

# Email Configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'chrisbenny2201@gmail.com'  # Change this
app.config['MAIL_PASSWORD'] = 'msgb qzil uynt zcwp'      # Use Gmail App Password
app.config['MAIL_DEFAULT_SENDER'] = 'Ticketing System <chrisbenny2201@gmail.com>'

mail = Mail(app)
db.init_app(app)

# ============= DATABASE HELPER FUNCTIONS =============

def get_all_users():
    """Get all users from database"""
    users = User.query.all()
    return {u.username: {
        'password': u.password,
        'role': u.role,
        'name': u.name,
        'can_assign': u.can_assign,
        'permissions': u.get_permissions()
    } for u in users}

def get_user_by_username(username):
    """Get single user from database"""
    user = db.session.get(User, username)
    if user:
        return {
            'password': user.password,
            'role': user.role,
            'name': user.name,
            'can_assign': user.can_assign,
            'permissions': user.get_permissions()
        }
    return None

def get_all_field_configs():
    """Get all field configs from database"""
    fields = FieldConfig.query.all()
    return {f.field_name: {
        'type': f.field_type,
        'label': f.label,
        'editable': f.editable,
        'options': f.get_options(),
        'permissions': f.get_permissions()
    } for f in fields}

def get_all_packages():
    """Get all packages from database"""
    packages = Package.query.all()
    return {p.id: {
        'id': p.id,
        'created_by': p.created_by,
        'created_at': p.created_at.isoformat() if p.created_at else None,
        'updated_by': p.updated_by,
        'updated_at': p.updated_at.isoformat() if p.updated_at else None,
        'assigned_users': p.get_assigned_users(),
        'gantt_columns': p.get_gantt_columns(),
        'data_timeline': p.get_data_timeline(),
        'data': p.get_data(),
        'user_permissions': p.get_user_permissions()  # NEW
    } for p in packages}

def can_user_perform_package_action(username, user_role, package, action):
    """
    Check if user can perform action on specific package
    action: 'read', 'update', 'delete'
    Returns: (can_perform, reason)
    """
    # Admin always can
    if user_role == 'admin':
        return True, "Admin access"
    
    # Check if assigned to package
    if username not in package.get_assigned_users():
        return False, "Not assigned to package"
    
    # Check package-specific permission (overrides global)
    pkg_user_perms = package.get_user_permissions()
    user_pkg_perms = pkg_user_perms.get(username, {})
    
    # Map action to permission key
    action_map = {
        'read': 'can_view',
        'update': 'can_edit',
        'delete': 'can_delete'
    }
    
    perm_key = action_map.get(action)
    if perm_key and perm_key in user_pkg_perms:
        # Package-specific permission exists, use it
        return user_pkg_perms[perm_key], f"Package-specific: {user_pkg_perms[perm_key]}"
    
    # Fall back to global user permissions
    # Note: This requires accessing user from database
    user = db.session.get(User, username)
    if not user:
        return False, "User not found"
    
    global_perms = user.get_permissions()
    global_perm_map = {
        'read': 'package_read',
        'update': 'package_update',
        'delete': 'package_delete'
    }
    
    global_key = global_perm_map.get(action)
    can_perform = global_perms.get(global_key, False)
    
    return can_perform, f"Global permission: {can_perform}"

def init_default_data():
    """Initialize default data in database"""
    with app.app_context():
        db.create_all()
        
        # Create default admin user if not exists
        if not db.session.get(User, 'admin'):
            admin = User(
                username='admin',
                password='admin123',
                role='admin',
                name='Administrator',
                can_assign=True
            )
            admin.set_permissions({
                'package_create': True,
                'package_read': True,
                'package_update': True,
                'package_delete': True
            })
            db.session.add(admin)
        
        # Create default user if not exists
        if not db.session.get(User, 'user1'):
            user1 = User(
                username='user1',
                password='user123',
                role='user',
                name='User One',
                can_assign=False
            )
            user1.set_permissions({
                'package_create': False,
                'package_read': True,
                'package_update': True,
                'package_delete': False
            })
            db.session.add(user1)
        
        db.session.commit()
        print("Default data initialized in database")

def login_required(f):
    """Decorator to require login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator to require admin role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        user = db.session.get(User, session['username'])
        if not user or user.role != 'admin':
            flash('Admin access required')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def permission_required(permission_name):
    """Decorator to check if user has specific permission"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'username' not in session:
                return redirect(url_for('login'))
            
            user = db.session.get(User, session['username'])
            if not user:
                return redirect(url_for('login'))
            
            # Admin always has all permissions
            if user.role == 'admin':
                return f(*args, **kwargs)
            
            # Check specific permission
            permissions = user.get_permissions()
            if not permissions.get(permission_name, False):
                flash(f'Permission denied: You cannot {permission_name.replace("_", " ")}')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.context_processor
def inject_fresh_permissions():
    """Inject fresh permissions and field access checker into all templates"""
    if 'username' in session:
        user = db.session.get(User, session['username'])
        if not user:
            return {'fresh_role': None, 'fresh_permissions': {}, 'can_access_field': lambda f, a: False}
        
        username = user.username
        user_role = user.role
        config = {'fields': get_all_field_configs()}
        
        def can_access_field(field_name, action='read'):
            """Check if current user can perform action on field"""
            # Admin always has full access
            if user_role == 'admin':
                return True
            
            field_config = config['fields'].get(field_name, {})
            field_permissions = field_config.get('permissions', {})
            
            # Check username-specific permissions first
            user_permissions = field_permissions.get(username, {})
            if user_permissions:
                return user_permissions.get(action, False)
            
            # Fallback to role-based if no username-specific permissions exist
            role_permissions = field_permissions.get(user_role, {})
            return role_permissions.get(action, False)
        
        return {
            'fresh_role': user_role,
            'fresh_permissions': user.get_permissions(),
            'can_access_field': can_access_field
        }
    return {
        'fresh_role': None,
        'fresh_permissions': {},
        'can_access_field': lambda f, a: False
    }

def send_async_email(app, msg):
    """Send email in background thread"""
    with app.app_context():
        try:
            mail.send(msg)
            print(f"Email sent successfully to {msg.recipients}")
        except Exception as e:
            print(f"Email send failed: {e}")

def send_professional_email(recipients, subject, template_name, **template_vars):
    """
    Send professional HTML email using templates
    recipients: list of email addresses
    template_name: e.g., 'package_created', 'package_updated'
    template_vars: variables to pass to template
    """
    try:
        # Render HTML template
        html_body = render_template(f'emails/{template_name}.html', **template_vars)
        
        # Create message
        msg = Message(
            subject=subject,
            recipients=recipients,
            html=html_body
        )
        
        # Send asynchronously
        Thread(target=send_async_email, args=(app, msg)).start()
        print(f"✉️ Email sent: {subject} to {recipients}")
        
    except Exception as e:
        print(f"❌ Email failed: {e}")


def notify_admins_package_change(package_id, changed_by, action_type, changes_detail, package_data=None):
    """
    Notify admins about package changes with professional HTML emails
    action_type: 'created', 'updated', 'deleted', 'column_added', 'column_deleted'
    """
    try:
        # Get all admin users with email preferences
        admins = User.query.filter_by(role='admin').all()
        
        for admin in admins:
            # Skip the admin who made the change
            if admin.username == changed_by:
                continue
            
            # Check email preferences
            prefs = db.session.get(EmailPreference, admin.username)
            
            if prefs and prefs.unsubscribed:
                continue
            
            # Check if this notification type is enabled
            if prefs:
                if action_type == 'created' and not prefs.package_created:
                    continue
                elif action_type == 'updated' and not prefs.package_updated:
                    continue
                elif action_type == 'deleted' and not prefs.package_deleted:
                    continue
                elif action_type in ['column_added', 'column_deleted'] and not prefs.column_added:
                    continue
            
            # Get email address
            email = prefs.email_address if prefs and prefs.email_address else 'chrisbenny2201@gmail.com'
            
            # Prepare template variables based on action type
            if action_type == 'created':
                subject = f"📦 New Package Created: {package_id}"
                template = 'package_created'
                template_vars = {
                    'admin_name': admin.name,
                    'package_id': package_id,
                    'created_by': changed_by,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'assigned_users': package_data.get('assigned_users', []) if package_data else [],
                    'status': package_data.get('status', 'N/A') if package_data else 'N/A',
                    'package_details': changes_detail,
                    'package_url': f"http://localhost:5000/packages/{package_id}/edit"
                }
            
            elif action_type == 'updated':
                subject = f"🔄 Package Updated: {package_id}"
                template = 'package_updated'
                changes_list = changes_detail.split('\n') if isinstance(changes_detail, str) else [changes_detail]
                template_vars = {
                    'admin_name': admin.name,
                    'package_id': package_id,
                    'updated_by': changed_by,
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'action_type': action_type.replace('_', ' ').title(),
                    'changes': [c for c in changes_list if c.strip()],
                    'package_url': f"http://localhost:5000/packages/{package_id}/edit"
                }
            
            elif action_type == 'deleted':
                subject = f"🗑️ Package Deleted: {package_id}"
                template = 'package_deleted'
                template_vars = {
                    'admin_name': admin.name,
                    'package_id': package_id,
                    'deleted_by': changed_by,
                    'deleted_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'reason': None,
                    'package_info': package_data or {}
                }
            
            elif action_type in ['column_added', 'column_deleted']:
                subject = f"📊 Package Modified: {package_id}"
                template = 'package_updated'
                template_vars = {
                    'admin_name': admin.name,
                    'package_id': package_id,
                    'updated_by': changed_by,
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'action_type': action_type.replace('_', ' ').title(),
                    'changes': [changes_detail],
                    'package_url': f"http://localhost:5000/packages/{package_id}/edit"
                }
            
            else:
                continue
            
            # Send email
            send_professional_email([email], subject, template, **template_vars)
    
    except Exception as e:
        print(f"❌ Notification error: {e}")

@app.route('/')
def index():
    if 'username' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = db.session.get(User, username)
        
        if user and user.password == password:
            session['username'] = username
            session['role'] = user.role
            session['name'] = user.name
            session['can_assign'] = user.can_assign
            session['permissions'] = user.get_permissions()
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid credentials')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    packages = get_all_packages()
    config = {'fields': get_all_field_configs()}
    users = get_all_users()  # ADD THIS LINE - needed for filter dropdown
    
    # Get filter parameter from URL
    filter_user = request.args.get('filter_user', 'all')
    
    # Filter packages based on user assignments
    if session['role'] != 'admin':
        # Non-admin users: only see their assigned packages
        user_packages = {}
        for pkg_id, pkg_data in packages.items():
            if session['username'] in pkg_data.get('assigned_users', []):
                user_packages[pkg_id] = pkg_data
        packages = user_packages
    else:
        # Admin users: can filter by specific user
        if filter_user != 'all':
            filtered_packages = {}
            for pkg_id, pkg_data in packages.items():
                if filter_user in pkg_data.get('assigned_users', []):
                    filtered_packages[pkg_id] = pkg_data
            packages = filtered_packages
    
    return render_template('dashboard.html', packages=packages, config=config, users=users)

@app.route('/charts-visualization')
@login_required
def charts_visualization():
    """Charts and Visualization page"""
    return render_template('charts_visualization.html')

@app.route('/admin')
@admin_required
def admin_panel():
    return render_template('admin.html')

@app.route('/admin/users')
@admin_required
def admin_users():
    users = get_all_users()
    return render_template('admin_users.html', users=users)

@app.route('/admin/users/create', methods=['POST'])
@admin_required
def create_user():
    username = request.form['username']
    password = request.form['password']
    role = request.form['role']
    name = request.form['name']
    can_assign = 'can_assign' in request.form
    
    if db.session.get(User, username):
        flash('Username already exists')
    else:
        user = User(
            username=username,
            password=password,
            role=role,
            name=name,
            can_assign=can_assign
        )
        user.set_permissions({
            'package_create': False,
            'package_read': True,
            'package_update': True,
            'package_delete': False
        })
        db.session.add(user)
        db.session.commit()
        flash('User created successfully')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/users/edit/<username>', methods=['GET', 'POST'])
@admin_required
def edit_user(username):
    user = db.session.get(User, username)
    
    if not user:
        flash('User not found')
        return redirect(url_for('admin_users'))
    
    if request.method == 'POST':
        # Don't allow changing admin username
        if username == 'admin' and request.form['username'] != 'admin':
            flash('Cannot change admin username')
            return redirect(url_for('edit_user', username=username))
        
        new_username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        name = request.form['name']
        can_assign = 'can_assign' in request.form
        
        # If username changed, create new user and delete old
        if new_username != username:
            if db.session.get(User, new_username):
                flash('Username already exists')
                return redirect(url_for('edit_user', username=username))
            
            # Create new user with same data
            new_user = User(
                username=new_username,
                password=password,
                role=role,
                name=name,
                can_assign=can_assign
            )
            new_user.set_permissions(user.get_permissions())
            db.session.add(new_user)
            db.session.delete(user)
        else:
            # Update existing user
            user.password = password
            user.role = role
            user.name = name
            user.can_assign = can_assign
        
        db.session.commit()
        flash('User updated successfully')
        return redirect(url_for('admin_users'))
    
    user_data = {
        'password': user.password,
        'role': user.role,
        'name': user.name,
        'can_assign': user.can_assign,
        'permissions': user.get_permissions()
    }
    return render_template('edit_user.html', username=username, user_data=user_data)

@app.route('/admin/users/delete/<username>')
@admin_required
def delete_user(username):
    if username == 'admin':
        flash('Cannot delete admin user')
    else:
        user = db.session.get(User, username)
        if user:
            db.session.delete(user)
            db.session.commit()
            flash('User deleted successfully')
        else:
            flash('User not found')
    return redirect(url_for('admin_users'))

@app.route('/admin/config')
@admin_required
def admin_config():
    config = {'fields': get_all_field_configs()}
    return render_template('admin_config.html', config=config)

@app.route('/admin/config/field/edit/<field_name>', methods=['GET', 'POST'])
@admin_required
def edit_field_config(field_name):
    field = db.session.get(FieldConfig, field_name)
    
    if not field:
        flash('Field not found')
        return redirect(url_for('admin_config'))
    
    if request.method == 'POST':
        new_field_name = request.form['field_name']
        field_type = request.form['field_type']
        field_label = request.form['field_label']
        options = request.form.get('options', '').split(',') if request.form.get('options') else []
        
        # If field name changed, create new and delete old
        if new_field_name != field_name:
            if db.session.get(FieldConfig, new_field_name):
                flash('Field name already exists')
                return redirect(url_for('edit_field_config', field_name=field_name))
            
            new_field = FieldConfig(
                field_name=new_field_name,
                field_type=field_type,
                label=field_label,
                editable=field.editable
            )
            new_field.set_options([opt.strip() for opt in options if opt.strip()])
            new_field.set_permissions(field.get_permissions())
            db.session.add(new_field)
            db.session.delete(field)
        else:
            field.field_type = field_type
            field.label = field_label
            field.set_options([opt.strip() for opt in options if opt.strip()])
        
        db.session.commit()
        flash('Field configuration updated successfully')
        return redirect(url_for('admin_config'))
    
    field_config = {
        'type': field.field_type,
        'label': field.label,
        'editable': field.editable,
        'options': field.get_options(),
        'permissions': field.get_permissions()
    }
    return render_template('edit_field.html', field_name=field_name, field_config=field_config)

@app.route('/admin/config/field', methods=['POST'])
@admin_required
def update_field_config():
    field_name = request.form['field_name']
    field_type = request.form['field_type']
    field_label = request.form['field_label']
    editable = 'editable' in request.form
    options = request.form.get('options', '').split(',') if request.form.get('options') else []
    
    if db.session.get(FieldConfig, field_name):
        flash('Field already exists')
        return redirect(url_for('admin_config'))
    
    field = FieldConfig(
        field_name=field_name,
        field_type=field_type,
        label=field_label,
        editable=editable
    )
    field.set_options([opt.strip() for opt in options if opt.strip()])
    
    # Set default permissions - username-based for all users
    users = User.query.all()
    default_perms = {}
    
    for user in users:
        if user.role == 'admin':
            default_perms['admin'] = {'read': True, 'write': True, 'delete': True}
        else:
            # Each user gets individual permissions
            default_perms[user.username] = {'read': True, 'write': editable, 'delete': False}
    
    field.set_permissions(default_perms)
    
    db.session.add(field)
    db.session.commit()
    flash('Field configuration updated')
    return redirect(url_for('admin_config'))

@app.route('/admin/config/field/delete/<field_name>')
@admin_required
def delete_field_config(field_name):
    field = db.session.get(FieldConfig, field_name)
    if field:
        db.session.delete(field)
        db.session.commit()
        flash('Field deleted successfully')
    return redirect(url_for('admin_config'))

@app.route('/packages/create', methods=['GET', 'POST'])
@login_required
@permission_required('package_create')
def create_package():
    if request.method == 'POST':
        config = {'fields': get_all_field_configs()}
        
        # Generate package ID
        pkg_count = Package.query.count()
        package_id = f"PKG_{pkg_count + 1:04d}"
        
        # Get assigned users (admin can assign, users auto-assign to themselves)
        if session['role'] == 'admin':
            assigned_users = request.form.getlist('assigned_users')
            if not assigned_users:
                flash('Please assign at least one user')
                return redirect(url_for('create_package'))
        else:
            # Non-admin users: auto-assign to themselves
            assigned_users = [session['username']]
        
        # Build package data with initial column
        initial_column_date = datetime.now().strftime('%Y-%m-%d')
        
        data_timeline = {initial_column_date: {}}
        
        # Add field data for initial column
        for field_name, field_config in config['fields'].items():
            if field_config['type'] == 'multiselect':
                data_timeline[initial_column_date][field_name] = request.form.getlist(field_name)
            else:
                data_timeline[initial_column_date][field_name] = request.form.get(field_name, '')
        
        package = Package(
            id=package_id,
            created_by=session['username'],
            created_at=datetime.now()
        )
        package.set_assigned_users(assigned_users)
        package.set_gantt_columns([initial_column_date])
        package.set_data_timeline(data_timeline)
        package.set_data(data_timeline[initial_column_date])
        
        db.session.add(package)
        db.session.commit()
        
        # Send notification (only if non-admin created package)
        if session.get('role') != 'admin':
            changes_detail = f"New package created\nAssigned to: {', '.join(assigned_users)}"
            notify_admins_package_change(
                package_id=package_id,
                changed_by=session['username'],
                action_type='created',
                changes_detail=changes_detail
            )
        
        flash('Package created successfully')
        return redirect(url_for('dashboard'))
    
    config = {'fields': get_all_field_configs()}
    users = get_all_users()
    return render_template('create_package.html', config=config, users=users)

@app.route('/packages/<package_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_package(package_id):
    package = db.session.get(Package, package_id)
    
    if not package:
        flash('Package not found')
        return redirect(url_for('dashboard'))
    
    user = db.session.get(User, session['username'])
    user_role = user.role
    username = session['username']
    
    # Check package-specific permissions
    can_edit, reason = can_user_perform_package_action(username, user_role, package, 'update')
    
    if not can_edit:
        flash(f'Access denied: {reason}')
        return redirect(url_for('dashboard'))
    
    config = {'fields': get_all_field_configs()}
    
    # Initialize gantt structure if not exists (for legacy packages)
    gantt_columns = package.get_gantt_columns()
    data_timeline = package.get_data_timeline()
    
    if not gantt_columns:
        initial_date = datetime.now().strftime('%Y-%m-%d')
        gantt_columns = [initial_date]
        data_timeline = {initial_date: package.get_data()}
        package.set_gantt_columns(gantt_columns)
        package.set_data_timeline(data_timeline)
    
    if request.method == 'POST':
        action = request.form.get('action', 'update')
        
        # Capture old state for comparison
        old_timeline = data_timeline.copy()
        old_assigned = set(package.get_assigned_users())
        
        if action == 'add_column':
            # Add new column
            new_date = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            gantt_columns.append(new_date)
            
            # Initialize new column data
            data_timeline[new_date] = {}
            
            # Copy from latest column
            latest_column = gantt_columns[-2] if len(gantt_columns) > 1 else gantt_columns[0]
            latest_data = data_timeline[latest_column]
            
            for field_name, field_config in config['fields'].items():
                field_permissions = field_config.get('permissions', {})
                
                # Check username-specific permissions first
                user_permissions = field_permissions.get(username, {})
                if not user_permissions:
                    # Fallback to role-based if username not found
                    user_permissions = field_permissions.get(user_role, {})
                
                can_read = user_role == 'admin' or user_permissions.get('read', False)
                
                if not can_read:
                    continue
                
                can_write = user_role == 'admin' or user_permissions.get('write', False)
                
                if not can_write:
                    data_timeline[new_date][field_name] = latest_data.get(field_name, '')
                else:
                    if field_config['type'] == 'multiselect':
                        data_timeline[new_date][field_name] = []
                    else:
                        data_timeline[new_date][field_name] = ''
            
            changes_detail = f"New column added: {new_date}"
            
        else:
            # Track changes
            changes = []
            
            # Check user assignments
            new_assigned = set(request.form.getlist('assigned_users'))
            if new_assigned != old_assigned:
                added = new_assigned - old_assigned
                removed = old_assigned - new_assigned
                if added:
                    changes.append(f"Users added: {', '.join(added)}")
                if removed:
                    changes.append(f"Users removed: {', '.join(removed)}")
            
            # Update existing columns
            for column_date in gantt_columns:
                for field_name, field_config in config['fields'].items():
                    field_permissions = field_config.get('permissions', {})
                    
                    # Check username-specific permissions first
                    user_permissions = field_permissions.get(username, {})
                    if not user_permissions:
                        # Fallback to role-based
                        user_permissions = field_permissions.get(user_role, {})
                    
                    can_write = user_role == 'admin' or user_permissions.get('write', False)
                    
                    if can_write:
                        form_field = f"{field_name}_{column_date}"
                        old_val = old_timeline.get(column_date, {}).get(field_name, '')
                        
                        if field_config['type'] == 'multiselect':
                            new_val = request.form.getlist(form_field)
                            data_timeline[column_date][field_name] = new_val
                        else:
                            new_val = request.form.get(form_field, '')
                            data_timeline[column_date][field_name] = new_val
                        
                        # Track change
                        if isinstance(old_val, list):
                            old_val_str = ', '.join(old_val)
                        else:
                            old_val_str = old_val
                        
                        if isinstance(new_val, list):
                            new_val_str = ', '.join(new_val)
                        else:
                            new_val_str = new_val
                        
                        if old_val_str != new_val_str:
                            changes.append(
                                f"- {field_config['label']} [{column_date}]:\n"
                                f"  Old: {old_val_str or '(empty)'}\n"
                                f"  New: {new_val_str or '(empty)'}"
                            )
            
            # Admin can update package-specific permissions
            if user_role == 'admin':
                # Update assigned users
                package.set_assigned_users(request.form.getlist('assigned_users'))
                
                # Update package-specific permissions for each assigned user
                pkg_perms = {}
                for assigned_user in request.form.getlist('assigned_users'):
                    user_pkg_perm = {}
                    
                    if f'can_edit_{assigned_user}' in request.form:
                        user_pkg_perm['can_edit'] = True
                    
                    if f'can_delete_{assigned_user}' in request.form:
                        user_pkg_perm['can_delete'] = True
                    
                    if user_pkg_perm:
                        pkg_perms[assigned_user] = user_pkg_perm

                package.set_user_permissions(pkg_perms)
            
            changes_detail = "\n".join(changes) if changes else "Minor updates (no field changes detected)"
        
        # Update package
        package.set_gantt_columns(gantt_columns)
        package.set_data_timeline(data_timeline)
        
        # Update legacy data field with latest column data
        if gantt_columns:
            latest_column = gantt_columns[-1]
            package.set_data(data_timeline[latest_column])
        
        package.updated_at = datetime.now()
        package.updated_by = username
        
        db.session.commit()
        
        # Send notification (only if non-admin made changes)
        if user_role != 'admin':
            notify_admins_package_change(
                package_id=package_id,
                changed_by=username,
                action_type=action if action == 'add_column' else 'updated',
                changes_detail=changes_detail
            )
        
        if action == 'add_column':
            flash('New column added successfully')
        else:
            flash('Package updated successfully')
        
        return redirect(url_for('edit_package', package_id=package_id))
    
    # Prepare package data for template
    package_data = {
        'id': package.id,
        'created_by': package.created_by,
        'created_at': package.created_at.isoformat() if package.created_at else None,
        'updated_by': package.updated_by,
        'updated_at': package.updated_at.isoformat() if package.updated_at else None,
        'assigned_users': package.get_assigned_users(),
        'user_permissions': package.get_user_permissions(),
        'gantt_columns': gantt_columns,
        'data_timeline': data_timeline,
        'data': package.get_data()
    }
    
    users = get_all_users()
    return render_template('edit_package.html', package=package_data, config=config, users=users)
@app.route('/packages/<package_id>/delete_column/<column_date>')
@login_required
def delete_column(package_id, column_date):
    package = db.session.get(Package, package_id)
    
    if not package:
        flash('Package not found')
        return redirect(url_for('dashboard'))
    
    user = db.session.get(User, session['username'])
    user_role = user.role
    
    # Check package-level access
    if user_role != 'admin' and session['username'] not in package.get_assigned_users():
        flash('Access denied')
        return redirect(url_for('dashboard'))
    
    config = {'fields': get_all_field_configs()}
    

    # Check if user has delete permission for ALL fields
    if user_role != 'admin':
        can_delete_column = True
        for field_name, field_config in config['fields'].items():
            field_permissions = field_config.get('permissions', {})
            
            # NEW: Check username-specific permissions first
            user_permissions = field_permissions.get(session['username'], {})
            if not user_permissions:
                # Fallback to role-based
                user_permissions = field_permissions.get(user_role, {})
            
            if not user_permissions.get('delete', False):
                can_delete_column = False
                break
        
        if not can_delete_column:
            flash('Permission denied: You cannot delete columns')
            return redirect(url_for('edit_package', package_id=package_id))
    
    gantt_columns = package.get_gantt_columns()
    
    # Don't allow deleting if only one column
    if len(gantt_columns) <= 1:
        flash('Cannot delete the last column')
        return redirect(url_for('edit_package', package_id=package_id))
    
    # Remove column
    if column_date in gantt_columns:
        gantt_columns.remove(column_date)
        data_timeline = package.get_data_timeline()
        if column_date in data_timeline:
            del data_timeline[column_date]
        
        package.set_gantt_columns(gantt_columns)
        package.set_data_timeline(data_timeline)
        
        # Update legacy data with latest column
        if gantt_columns:
            latest_column = gantt_columns[-1]
            package.set_data(data_timeline[latest_column])
        
        db.session.commit()
        if user_role != 'admin':
            notify_admins_package_change(
                package_id=package_id,
                changed_by=session['username'],
                action_type='column_deleted',
                changes_detail=f"Column deleted: {column_date}"
            )
        flash('Column deleted successfully')
    
    return redirect(url_for('edit_package', package_id=package_id))

@app.route('/packages/<package_id>/delete')
@login_required
def delete_package(package_id):
    package = db.session.get(Package, package_id)
    
    if not package:
        flash('Package not found')
        return redirect(url_for('dashboard'))
    
    user_role = session['role']
    username = session['username']
    
    # Check package-specific delete permission
    can_delete, reason = can_user_perform_package_action(username, user_role, package, 'delete')
    
    if not can_delete:
        flash(f'Permission denied: {reason}')
        return redirect(url_for('dashboard'))
    
    # Send notification (only if non-admin deleted package)
    if user_role != 'admin':
        notify_admins_package_change(
            package_id=package_id,
            changed_by=username,
            action_type='deleted',
            changes_detail=f"Package {package_id} was deleted"
        )
    
    db.session.delete(package)
    db.session.commit()
    flash('Package deleted successfully')
    return redirect(url_for('dashboard'))

@app.route('/admin/permissions', methods=['GET', 'POST'])
@admin_required
def admin_permissions():
    users = User.query.all()
    
    if request.method == 'POST':
        # Update permissions for each user
        for user in users:
            if user.username != 'admin':
                new_permissions = {
                    'package_create': f'create_{user.username}' in request.form,
                    'package_read': f'read_{user.username}' in request.form,
                    'package_update': f'update_{user.username}' in request.form,
                    'package_delete': f'delete_{user.username}' in request.form
                }
                user.set_permissions(new_permissions)
                
                # THIS IS THE KEY: Update session if user is logged in elsewhere
                # Check all active sessions (this is simplified - works for single user)
                if user.username == session.get('username'):
                    session['permissions'] = new_permissions
                    session.modified = True
        
        db.session.commit()
        flash('Permissions updated successfully. Changes apply immediately.')
        return redirect(url_for('admin_permissions'))
    
    users_dict = get_all_users()
    return render_template('admin_permissions.html', users=users_dict)

@app.route('/admin/field-permissions/<field_name>', methods=['GET', 'POST'])
@admin_required
def admin_field_permissions(field_name):
    field = db.session.get(FieldConfig, field_name)
    
    if not field:
        flash('Field not found')
        return redirect(url_for('admin_config'))
    
    users = User.query.all()
    
    if request.method == 'POST':
        permissions = {}
        
        for user in users:
            if user.role == 'admin':
                # Admin always has all permissions
                permissions['admin'] = {'read': True, 'write': True, 'delete': True}
            else:
                # Individual username-based permissions
                permissions[user.username] = {
                    'read': f'read_{user.username}' in request.form,
                    'write': f'write_{user.username}' in request.form,
                    'delete': f'delete_{user.username}' in request.form
                }
        
        field.set_permissions(permissions)
        db.session.commit()
        flash(f'Permissions updated for field: {field_name}')
        return redirect(url_for('admin_config'))
    
    field_config = {
        'type': field.field_type,
        'label': field.label,
        'editable': field.editable,
        'options': field.get_options(),
        'permissions': field.get_permissions()
    }
    current_permissions = field.get_permissions()
    
    return render_template('admin_field_permissions.html', 
                         field_name=field_name,
                         field_config=field_config,
                         users=users,
                         current_permissions=current_permissions)

def migrate_field_permissions_to_username():
    """Migrate role-based field permissions to username-based"""
    with app.app_context():
        fields = FieldConfig.query.all()
        users = User.query.all()
        
        for field in fields:
            current_perms = field.get_permissions()
            new_perms = {}
            
            # Keep admin role-based
            if 'admin' in current_perms:
                new_perms['admin'] = current_perms['admin']
            else:
                new_perms['admin'] = {'read': True, 'write': True, 'delete': True}
            
            # Convert each user
            for user in users:
                if user.role != 'admin':
                    # Check if username already exists
                    if user.username in current_perms:
                        new_perms[user.username] = current_perms[user.username]
                    # Otherwise use role-based as template
                    elif user.role in current_perms:
                        new_perms[user.username] = current_perms[user.role].copy()
                    # Default permissions
                    else:
                        new_perms[user.username] = {'read': True, 'write': field.editable, 'delete': False}
            
            field.set_permissions(new_perms)
        
        db.session.commit()
        print("Field permissions migrated to username-based successfully!")

@app.route('/export/excel')
@login_required
def export_excel():
    """Export packages to Excel"""
    packages = get_all_packages()
    config = {'fields': get_all_field_configs()}

    filter_user = request.args.get('filter_user', 'all')
    
    # Filter packages based on user role
    if session['role'] != 'admin':
        user_packages = {}
        for pkg_id, pkg_data in packages.items():
            if session['username'] in pkg_data.get('assigned_users', []):
                user_packages[pkg_id] = pkg_data
        packages = user_packages
    else:
        # Admin users: can filter by specific user
        if filter_user != 'all':
            filtered_packages = {}
            for pkg_id, pkg_data in packages.items():
                if filter_user in pkg_data.get('assigned_users', []):
                    filtered_packages[pkg_id] = pkg_data
            packages = filtered_packages
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Packages"
    
    # Define headers
    headers = ['Package ID', 'Status', 'Assigned Users', 'Created By', 'Created Date']
    
    # Add custom field headers
    for field_name, field_config in config['fields'].items():
        if field_name.upper() not in ['STATUS']:
            headers.append(field_config['label'])
    
    # Style for headers
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    row_num = 2
    for package_id, package in packages.items():
        # Get latest data
        if package.get('data_timeline'):
            latest_column = package['gantt_columns'][-1]
            latest_data = package['data_timeline'][latest_column]
        else:
            latest_data = package.get('data', {})
        
        # Package ID
        ws.cell(row=row_num, column=1, value=package_id).border = thin_border
        
        # Status
        status_value = latest_data.get('STATUS', latest_data.get('status', '-'))
        status_cell = ws.cell(row=row_num, column=2, value=status_value)
        status_cell.border = thin_border
        
        # Color code status
        if str(status_value).lower() == 'ongoing':
            status_cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)
        elif str(status_value).lower() == 'hold':
            status_cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
            status_cell.font = Font(bold=True)
        elif str(status_value).lower() == 'completed':
            status_cell.fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)
        
        # Assigned Users
        assigned = ', '.join(package.get('assigned_users', []))
        ws.cell(row=row_num, column=3, value=assigned).border = thin_border
        
        # Created By
        ws.cell(row=row_num, column=4, value=package.get('created_by', '-')).border = thin_border
        
        # Created Date
        created_at = package.get('created_at', '-')
        if created_at and created_at != '-':
            created_at = created_at[:10]
        ws.cell(row=row_num, column=5, value=created_at).border = thin_border
        
        # Custom fields
        col_num = 6
        for field_name, field_config in config['fields'].items():
            if field_name.upper() not in ['STATUS']:
                field_value = latest_data.get(field_name, '-')
                
                # Handle multiselect
                if isinstance(field_value, list):
                    field_value = ', '.join(field_value)
                
                ws.cell(row=row_num, column=col_num, value=field_value).border = thin_border
                col_num += 1
        
        row_num += 1
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"packages_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/email-preferences', methods=['GET', 'POST'])
@login_required
@admin_required
def email_preferences():
    """Manage email notification preferences"""
    username = session['username']
    prefs = db.session.get(EmailPreference, username)
    
    # Create default preferences if not exist
    if not prefs:
        prefs = EmailPreference(username=username)
        db.session.add(prefs)
        db.session.commit()
    
    if request.method == 'POST':
        # Update preferences
        prefs.package_created = 'package_created' in request.form
        prefs.package_updated = 'package_updated' in request.form
        prefs.package_deleted = 'package_deleted' in request.form
        prefs.column_added = 'column_added' in request.form
        prefs.column_deleted = 'column_deleted' in request.form
        
        prefs.email_address = request.form.get('email_address', '')
        
        db.session.commit()
        flash('Email preferences updated successfully')
        return redirect(url_for('email_preferences'))
    
    return render_template('email_preferences.html', prefs=prefs.get_preferences())

@app.route('/api/dashboard-stats')
@login_required
def dashboard_stats():
    """API endpoint to get dashboard statistics for charts"""
    packages = get_all_packages()
    config = {'fields': get_all_field_configs()}
    
    # Filter packages based on user role
    if session['role'] != 'admin':
        user_packages = {}
        for pkg_id, pkg_data in packages.items():
            if session['username'] in pkg_data.get('assigned_users', []):
                user_packages[pkg_id] = pkg_data
        packages = user_packages
    
    # Calculate statistics
    stats = {
        'total': len(packages),
        'ongoing': 0,
        'completed': 0,
        'hold': 0,
        'cancelled': 0,
        'not_started': 0,
        'delayed': 0,
        'packages_per_user': {},
        'monthly_trend': {}
    }
    
    today = datetime.now().date()
    
    for pkg_id, pkg in packages.items():
        # Get latest status
        if pkg.get('data_timeline'):
            latest_col = pkg['gantt_columns'][-1]
            latest_data = pkg['data_timeline'][latest_col]
        else:
            latest_data = pkg.get('data', {})
        
        status = latest_data.get('STATUS', latest_data.get('status', '')).lower()
        sow = latest_data.get('SOW', '')
        sow_date = latest_data.get('SOW DATE', '')
        actual_closure = latest_data.get('ACTUAL CLOSURE DATE', '')
        commercial_target = latest_data.get('COMMERCIAL TARGET DATE', '')
        
        # Count by status
        if actual_closure:
            stats['completed'] += 1
        elif not sow or sow.upper() == 'NO' or not sow_date:
            stats['not_started'] += 1
        elif status == 'ongoing':
            stats['ongoing'] += 1
        elif status == 'hold':
            stats['hold'] += 1
        elif status == 'cancelled':
            stats['cancelled'] += 1
        
        # Check if delayed
        if commercial_target and not actual_closure:
            try:
                target_date = datetime.strptime(commercial_target, '%Y-%m-%d').date()
                if target_date < today:
                    stats['delayed'] += 1
            except:
                pass
        
        # Count per user
        for user in pkg.get('assigned_users', []):
            stats['packages_per_user'][user] = stats['packages_per_user'].get(user, 0) + 1
        
        # Monthly trend
        created_at = pkg.get('created_at')
        if created_at:
            try:
                month = created_at[:7]  # YYYY-MM
                stats['monthly_trend'][month] = stats['monthly_trend'].get(month, 0) + 1
            except:
                pass
    
    return jsonify(stats)

@app.route('/api/gantt-data')
@login_required
def gantt_data():
    """API endpoint to get Gantt chart data"""
    packages = get_all_packages()
    config = {'fields': get_all_field_configs()}
    
    # Filter packages based on user role
    if session['role'] != 'admin':
        user_packages = {}
        for pkg_id, pkg_data in packages.items():
            if session['username'] in pkg_data.get('assigned_users', []):
                user_packages[pkg_id] = pkg_data
        packages = user_packages
    
    # Build Gantt data
    gantt_packages = []
    today = datetime.now().date()
    
    for pkg_id, pkg in packages.items():
        # Get latest data
        if pkg.get('data_timeline'):
            latest_col = pkg['gantt_columns'][-1]
            latest_data = pkg['data_timeline'][latest_col]
        else:
            latest_data = pkg.get('data', {})
        
        sow = latest_data.get('SOW', '')
        sow_date = latest_data.get('SOW DATE', '')
        commercial_target = latest_data.get('COMMERCIAL TARGET DATE', '')
        actual_closure = latest_data.get('ACTUAL CLOSURE DATE', '')
        created_at = pkg.get('created_at', '')
        
        # Determine status
        status = 'not_started'
        
        # Check if package has started
        if sow and sow.upper() != 'NO' and sow_date:
            status = 'ongoing'
            
            # Check if completed
            if actual_closure:
                status = 'completed'
            # Check if overdue
            elif commercial_target:
                try:
                    target_date = datetime.strptime(commercial_target, '%Y-%m-%d').date()
                    if target_date < today:
                        status = 'overdue'
                except:
                    pass
        
        # Determine start and end dates for Gantt display
        start_date = None
        end_date = None
        
        # Priority 1: Use actual dates if available
        if sow_date:
            start_date = sow_date
        elif created_at:
            # Fallback: use creation date as start
            try:
                start_date = created_at[:10]  # Extract YYYY-MM-DD
            except:
                start_date = today.strftime('%Y-%m-%d')
        else:
            # Last resort: use today as start
            start_date = today.strftime('%Y-%m-%d')
        
        # Determine end date
        if actual_closure:
            end_date = actual_closure
        elif commercial_target:
            end_date = commercial_target
        else:
            # Fallback: use start + 30 days or today (whichever is later)
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
                default_end = start_dt + timedelta(days=30)
                end_date = max(default_end, today).strftime('%Y-%m-%d')
            except:
                end_date = today.strftime('%Y-%m-%d')
        
        # Ensure end date is not before start date
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            if end_dt < start_dt:
                end_date = (start_dt + timedelta(days=30)).strftime('%Y-%m-%d')
        except:
            pass
        
        gantt_packages.append({
            'id': pkg_id,
            'start': start_date,
            'end': end_date,
            'status': status,
            'assigned_users': pkg.get('assigned_users', [])
        })
    
    return jsonify(gantt_packages)
if __name__ == '__main__':
    init_default_data()
    migrate_field_permissions_to_username()
    app.run(debug=True)